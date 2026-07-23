# Verify every code snippet shown in book/ch14.js.
# Same self-checking idea as ch2_verify.py: each snippet's expected output
# (as embedded in the chapter file) is asserted here, so any drift between
# the chapter text and real execution fails the run.
#
# The openpyxl snippets need the third-party openpyxl module (the chapter
# discloses it runs on desktop Python, not in the app). When openpyxl is not
# installed, that whole section is skipped WITHOUT failing; the stdlib checks
# — including all four graded-exercise reference solutions — always run.
import io, contextlib, os, sys, tempfile

fails = 0
checks = 0
def check(label, got, want):
    global fails, checks
    checks += 1
    ok = got == want
    print(f"{'ok  ' if ok else 'FAIL'} {label}")
    if not ok:
        fails += 1
        print(f"     got:  {got!r}")
        print(f"     want: {want!r}")

def repl_seq(label, steps, expected, ns=None):
    # steps: list of (code, is_stmt) sharing one namespace.
    # expected: list of repr/error strings for the non-statement steps.
    if ns is None:
        ns = {}
    got = []
    for code, is_stmt in steps:
        if is_stmt:
            exec(code, ns)
        else:
            try:
                got.append(repr(eval(code, ns)))
            except Exception as ex:
                got.append(f"{type(ex).__name__}: {ex}")
    check(label, got, expected)

def run(code):
    buf = io.StringIO()
    ns = {}
    with contextlib.redirect_stdout(buf):
        exec(code, ns)
    return buf.getvalue().rstrip("\n")

print("=== stdlib: arithmetic behind the revenue project ===")
repl_seq("loaves times price per bread, and the order total",
    [("12 * 6.5", False), ("8 * 5.25", False), ("20 * 3.25", False),
     ("78.0 + 42.0 + 65.0", False)],
    ["78.0", "42.0", "65.0", "185.0"])

print("=== graded exercises: reference solutions vs expected o ===")
check("ex1 Total a loaf column",
    run("""rows = [['Item', 'Loaves'], ['Sourdough', 12], ['Rye', 8], ['Baguette', 20]]
total = 0
for row in rows[1:]:
    print(row[0] + ': ' + str(row[1]))
    total = total + row[1]
print('Total loaves: ' + str(total))"""),
    "Sourdough: 12\nRye: 8\nBaguette: 20\nTotal loaves: 40")
check("ex2 Rows into columns",
    run("""grid = [[1, 2, 3], [4, 5, 6]]
for c in range(3):
    column = []
    for r in range(2):
        column.append(grid[r][c])
    print(column)"""),
    "[1, 4]\n[2, 5]\n[3, 6]")
check("ex3 A sheet in a dictionary",
    run("""sheet = {'A1': 'Item', 'B1': 'Qty', 'A2': 'Nails', 'B2': 150, 'A3': 'Screws', 'B3': 95, 'A4': 'Hooks', 'B4': 60}
total = 0
for r in range(2, 5):
    qty = sheet['B' + str(r)]
    print(sheet['A' + str(r)] + ': ' + str(qty))
    total = total + qty
print('Total: ' + str(total))"""),
    "Nails: 150\nScrews: 95\nHooks: 60\nTotal: 305")
check("ex4 The price fixer",
    run("""rows = [['Rope', 40], ['Chalk', 21], ['Tape', 12], ['Harness', 55]]
fixes = {'Chalk': 9, 'Tape': 4}
for row in rows:
    if row[0] in fixes:
        row[1] = fixes[row[0]]
    print(row[0] + ': $' + str(row[1]))"""),
    "Rope: $40\nChalk: $9\nTape: $4\nHarness: $55")

# ---------------------------------------------------------------------------
# openpyxl-dependent checks. Guarded: skipping is not failing.
# ---------------------------------------------------------------------------
OPENPYXL_CHECKS = 24
try:
    import openpyxl
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

if not HAVE_OPENPYXL:
    print(f"skip: openpyxl not installed ({OPENPYXL_CHECKS} checks skipped)")
else:
    before = checks
    olddir = os.getcwd()
    with tempfile.TemporaryDirectory() as td:
        os.chdir(td)
        try:
            # Rebuild the same bakery.xlsx the chapter reads from.
            _wb = openpyxl.Workbook()
            _sh = _wb.active
            _sh.title = 'Orders'
            _data = [['Item', 'Loaves', 'Price'],
                     ['Sourdough', 12, 6.5],
                     ['Rye', 8, 5.25],
                     ['Baguette', 20, 3.25]]
            for _r in range(len(_data)):
                for _c in range(len(_data[_r])):
                    _sh.cell(row=_r + 1, column=_c + 1).value = _data[_r][_c]
            _sup = _wb.create_sheet(title='Suppliers')
            _sup['A1'] = 'Name'
            _sup['A2'] = 'Millbrook Flour'
            _sup['A3'] = 'Northfield Dairy'
            _wb.save('bakery.xlsx')

            print("=== section: Reading a workbook ===")
            ns2 = {}
            repl_seq("load_workbook type and sheetnames",
                [("import openpyxl", True),
                 ("wb = openpyxl.load_workbook('bakery.xlsx')", True),
                 ("type(wb)", False), ("wb.sheetnames", False)],
                ["<class 'openpyxl.workbook.workbook.Workbook'>",
                 "['Orders', 'Suppliers']"], ns2)
            repl_seq("sheet by name, active, missing sheet KeyError",
                [("sheet = wb['Orders']", True),
                 ("sheet.title", False), ("wb.active", False), ("wb['Nope']", False)],
                ["'Orders'", '<Worksheet "Orders">',
                 "KeyError: 'Worksheet Nope does not exist.'"], ns2)
            repl_seq("cell object vs value, coordinate parts",
                [("sheet['A1']", False), ("sheet['A1'].value", False),
                 ("c = sheet['B2']", True),
                 ("c.value", False), ("c.coordinate", False),
                 ("c.row", False), ("c.column_letter", False)],
                ["<Cell 'Orders'.A1>", "'Item'", "12", "'B2'", "2", "'B'"], ns2)
            repl_seq("cell(row, column) and sheet extent",
                [("sheet.cell(row=2, column=1)", False),
                 ("sheet.cell(row=2, column=1).value", False),
                 ("sheet.max_row", False), ("sheet.max_column", False)],
                ["<Cell 'Orders'.A2>", "'Sourdough'", "4", "3"], ns2)
            check("loop over column 2 by number",
                run("""import openpyxl
wb = openpyxl.load_workbook('bakery.xlsx')
sheet = wb['Orders']
for r in range(1, 5):
    print(sheet.cell(row=r, column=2).value)"""),
                "Loaves\n12\n8\n20")

            print("=== section: Column letters and slices ===")
            repl_seq("get_column_letter / column_index_from_string",
                [("from openpyxl.utils import get_column_letter, column_index_from_string", True),
                 ("get_column_letter(1)", False), ("get_column_letter(3)", False),
                 ("get_column_letter(27)", False),
                 ("get_column_letter(sheet.max_column)", False),
                 ("column_index_from_string('AA')", False)],
                ["'A'", "'C'", "'AA'", "'C'", "27"], ns2)
            check("slice loop A1:C2",
                run("""import openpyxl
wb = openpyxl.load_workbook('bakery.xlsx')
sheet = wb['Orders']
for row in sheet['A1':'C2']:
    for c in row:
        print(c.coordinate, c.value)
    print('--- row ends ---')"""),
                "A1 Item\nB1 Loaves\nC1 Price\n--- row ends ---\nA2 Sourdough\nB2 12\nC2 6.5\n--- row ends ---")
            check("whole column B",
                run("""import openpyxl
wb = openpyxl.load_workbook('bakery.xlsx')
sheet = wb['Orders']
for c in sheet['B']:
    print(c.value)"""),
                "Loaves\n12\n8\n20")

            print("=== section: the reading project ===")
            check("revenue project output",
                run("""import openpyxl
wb = openpyxl.load_workbook('bakery.xlsx')
sheet = wb['Orders']
revenue = {}
for r in range(2, sheet.max_row + 1):
    item = sheet.cell(row=r, column=1).value
    loaves = sheet.cell(row=r, column=2).value
    price = sheet.cell(row=r, column=3).value
    revenue[item] = loaves * price
total = 0
for item in revenue:
    print(item + ': $' + str(revenue[item]))
    total = total + revenue[item]
print('Whole order: $' + str(total))"""),
                "Sourdough: $78.0\nRye: $42.0\nBaguette: $65.0\nWhole order: $185.0")
            try:
                ns2['sheet'].cell(row=0, column=1)
                got = "no error"
            except Exception as ex:
                got = type(ex).__name__
            check("q5: row 0 raises an error (1-based indexes)", got, "ValueError")

            print("=== section: Writing a workbook ===")
            ns5 = {}
            repl_seq("new Workbook starts with one sheet named Sheet",
                [("import openpyxl", True),
                 ("wb = openpyxl.Workbook()", True),
                 ("wb.sheetnames", False),
                 ("sheet = wb.active", True),
                 ("sheet.title", False),
                 ("sheet.title = 'Stock'", True),
                 ("wb.sheetnames", False)],
                ["['Sheet']", "'Sheet'", "['Stock']"], ns5)
            repl_seq("create_sheet, index/title kwargs, del",
                [("wb.create_sheet()", False),
                 ("wb.create_sheet(index=0, title='Front')", False),
                 ("wb.sheetnames", False),
                 ("del wb['Sheet']", True),
                 ("wb.sheetnames", False)],
                ['<Worksheet "Sheet">', '<Worksheet "Front">',
                 "['Front', 'Stock', 'Sheet']", "['Front', 'Stock']"], ns5)
            repl_seq("writing a cell value",
                [("sheet['A1'] = 'Item'", True), ("sheet['A1'].value", False)],
                ["'Item'"], ns5)
            pre = os.path.exists('gear.xlsx')
            exec("wb.save('gear.xlsx')", ns5)
            post = os.path.exists('gear.xlsx')
            check("q7: nothing on disk until save()", [pre, post], [False, True])
            rows_before = openpyxl.load_workbook('gear.xlsx')['Stock'].max_row

            print("=== section: the writing project ===")
            check("builder program output",
                run("""import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Stock'
rows = [['Item', 'Price'], ['Rope', 40], ['Chalk', 21], ['Tape', 12], ['Harness', 55]]
for r in range(len(rows)):
    sheet.cell(row=r + 1, column=1).value = rows[r][0]
    sheet.cell(row=r + 1, column=2).value = rows[r][1]
wb.save('gear.xlsx')
print('saved gear.xlsx with', sheet.max_row, 'rows')"""),
                "saved gear.xlsx with 5 rows")
            rows_after = openpyxl.load_workbook('gear.xlsx')['Stock'].max_row
            check("q8: save() replaces an existing file silently",
                [rows_before, rows_after], [1, 5])
            check("fixer program output",
                run("""import openpyxl
FIXES = {'Chalk': 9, 'Tape': 4}
wb = openpyxl.load_workbook('gear.xlsx')
sheet = wb['Stock']
for r in range(2, sheet.max_row + 1):
    item = sheet.cell(row=r, column=1).value
    if item in FIXES:
        sheet.cell(row=r, column=2).value = FIXES[item]
wb.save('gear-fixed.xlsx')
wb2 = openpyxl.load_workbook('gear-fixed.xlsx')
sheet2 = wb2['Stock']
for r in range(2, sheet2.max_row + 1):
    print(sheet2.cell(row=r, column=1).value, sheet2.cell(row=r, column=2).value)"""),
                "Rope 40\nChalk 9\nTape 4\nHarness 55")

            print("=== section: Fonts and formulas ===")
            ns7 = {}
            repl_seq("Font assignment reads back (size becomes 18.0)",
                [("import openpyxl", True),
                 ("from openpyxl.styles import Font", True),
                 ("wb = openpyxl.Workbook()", True),
                 ("sheet = wb.active", True),
                 ("sheet['A1'] = 'Stock report'", True),
                 ("sheet['A1'].font = Font(size=18, bold=True)", True),
                 ("sheet['A1'].font.bold", False),
                 ("sheet['A1'].font.size", False)],
                ["True", "18.0"], ns7)
            repl_seq("formula cell stores the string",
                [("sheet['B2'] = 120", True),
                 ("sheet['B3'] = 45", True),
                 ("sheet['B4'] = '=SUM(B2:B3)'", True),
                 ("sheet['B4'].value", False),
                 ("wb.save('styled.xlsx')", True)],
                ["'=SUM(B2:B3)'"], ns7)
            check("formula read-back: string, and None under data_only",
                run("""import openpyxl
wb = openpyxl.load_workbook('styled.xlsx')
print(wb.active['B4'].value)
wb2 = openpyxl.load_workbook('styled.xlsx', data_only=True)
print(wb2.active['B4'].value)"""),
                "=SUM(B2:B3)\nNone")

            print("=== section: Layout ===")
            repl_seq("row height and column width round-trip",
                [("sheet.row_dimensions[1].height = 70", True),
                 ("sheet.row_dimensions[1].height", False),
                 ("sheet.column_dimensions['B'].width = 25", True),
                 ("sheet.column_dimensions['B'].width", False)],
                ["70.0", "25.0"], ns7)
            repl_seq("merge and unmerge",
                [("sheet.merge_cells('A6:D6')", True),
                 ("str(sheet.merged_cells)", False),
                 ("sheet.unmerge_cells('A6:D6')", True),
                 ("str(sheet.merged_cells)", False)],
                ["'A6:D6'", "''"], ns7)
            repl_seq("freeze_panes: 'A2', 'B2', and 'A1'/None meaning unfrozen",
                [("sheet.freeze_panes = 'A2'", True),
                 ("sheet.freeze_panes", False),
                 ("sheet.freeze_panes = 'B2'", True),
                 ("sheet.freeze_panes", False),
                 ("sheet.freeze_panes = 'A1'", True),
                 ("sheet.freeze_panes", False),
                 ("sheet.freeze_panes = None", True),
                 ("sheet.freeze_panes", False)],
                ["'A2'", "'B2'", "None", "None"], ns7)
            run("""import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
counts = [12, 8, 20, 15, 9]
for i in range(5):
    sheet.cell(row=i + 1, column=1).value = counts[i]
ref = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=5)
series = openpyxl.chart.Series(ref, title='Loaves')
chart = openpyxl.chart.BarChart()
chart.title = 'Loaves per day'
chart.append(series)
sheet.add_chart(chart, 'C2')
wb.save('chart.xlsx')""")
            check("chart program runs and saves chart.xlsx",
                os.path.exists('chart.xlsx'), True)
        finally:
            os.chdir(olddir)
    delta = checks - before
    check("guarded section ran the declared number of checks",
        delta, OPENPYXL_CHECKS)

print()
print("CH14 VERIFY: ALL PASS" if fails == 0 else f"CH14 VERIFY: {fails} FAILURE(S)")
sys.exit(0 if fails == 0 else 1)
